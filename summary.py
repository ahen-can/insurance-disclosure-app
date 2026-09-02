"""The '00 Summary' sheet: what was extracted, and what to go and re-check.

Two independent signals sit side by side here, and the difference between them
matters:

  Model confidence  what the model says about its own reading. Cheap, covers
                    every field, and unreliable in exactly the way you would
                    expect -- it cannot report that it read the wrong column,
                    because it does not know it did.

  Cross-checks      totals recomputed from the values that were written, by
                    checks.py. Narrow, but hard evidence: if a form's parts do
                    not add up to its own total, something was misread, and no
                    amount of model confidence changes that.

A form is only called clean when both agree. The verdict column is deliberately
pessimistic -- the cost of re-reading one page of a PDF is far below the cost of
a wrong figure reaching the repository.
"""

from datetime import datetime

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import checks

SHEET_NAME = "00 Summary"

INK = "FF2B2F33"
HEAD_FILL = PatternFill("solid", fgColor="FF17365D")
BAND_FILL = PatternFill("solid", fgColor="FFD9EAF7")
OK_FILL = PatternFill("solid", fgColor="FFE2EFDA")
REVIEW_FILL = PatternFill("solid", fgColor="FFFFF2CC")
CHECK_FILL = PatternFill("solid", fgColor="FFF8CBAD")

VERDICT_FILLS = {"OK": OK_FILL, "Review": REVIEW_FILL, "Check": CHECK_FILL}

FORM_HEADERS = ["Insurer", "Form", "Sheet", "Fields", "Filled", "Coverage",
                "Model confidence", "Cross-checks", "Verdict", "Note"]
REVIEW_HEADERS = ["Insurer", "Form", "Field", "Column", "Value", "Raised by",
                  "Reason"]
WIDTHS = [34, 8, 26, 8, 8, 10, 17, 22, 10, 60]


def _field_map(kind):
    if kind == "general":
        from config_general import FIELD_LABELS, SHEET_CONFIG_GENERAL, SHEET_TITLES
        return SHEET_CONFIG_GENERAL, SHEET_TITLES, FIELD_LABELS
    from config import FIELD_LABELS, SHEET_CONFIG
    titles = {code: code for code in SHEET_CONFIG}
    return SHEET_CONFIG, titles, FIELD_LABELS


def _confidence(data, form):
    """The model's score for one form, or None if it did not report one."""
    entry = (data.get("confidence") or {}).get(form)
    if isinstance(entry, dict):
        score, note = entry.get("score"), entry.get("note")
    else:
        score, note = entry, None
    return (float(score) if isinstance(score, (int, float)) else None), note


def _verdict(coverage, score, failed, flagged):
    """Coverage deliberately does not drive this.

    The prompt asks for 0 where a value is absent, so a legitimately nil line --
    no catastrophe reserve, no aviation book, nothing outside India -- is
    indistinguishable from one the model could not find. Letting a low count
    raise "Review" put 12 of 19 forms on the list for the first real filing and
    buried the three that had actually failed their arithmetic. Coverage stays
    on the sheet as a column to read; it is not evidence of an error.
    """
    if failed or (score is not None and score < 0.5) or coverage < 0.15:
        return "Check"
    if flagged or (score is not None and score < 0.8):
        return "Review"
    return "OK"


def collect(results: dict, kind: str, expected_year=None) -> tuple:
    """Build (form rows, review rows) without touching a workbook.

    Split out from the writing so it can be tested, and so the on-screen view
    can show the same numbers the sheet does.
    """
    config, titles, labels = _field_map(kind)
    form_rows, review_rows = [], []

    for insurer, result in results.items():
        if result.get("status") != "success":
            form_rows.append([insurer, "-", "-", "", "", "", "", "",
                              "Check", f"Extraction failed: {result.get('error', '')}"])
            continue

        data = result.get("data", {})
        model_review = data.get("review") or []

        # The general template is laid out for one financial year. Filing a
        # different year into it is silent and irreversible once the workbook
        # is keyed into the repository, so it is flagged loudly rather than
        # blocked -- the extraction may still be the one that was wanted.
        found_year = data.get("year")
        if expected_year and found_year and found_year != expected_year:
            review_rows.append([
                insurer, "-", "year", "", found_year, "Year check",
                f"This template is laid out for {expected_year} but the PDF was "
                f"read as {found_year}. Either the wrong year's disclosure was "
                f"uploaded, or the wrong column was read.",
            ])

        for form, fields in config.items():
            values = data.get(form)
            if not isinstance(values, dict):
                form_rows.append([insurer, form, titles.get(form, form), len(fields),
                                  0, 0.0, "not returned", "", "Check",
                                  "The model returned nothing for this form"])
                continue

            filled = sum(1 for f in fields
                         if isinstance(values.get(f), (int, float)) and values.get(f))
            coverage = filled / len(fields) if fields else 0.0
            score, note = _confidence(data, form)

            outcomes = checks.run_checks(form, values, kind)
            failed = [c for c in outcomes if not c["passed"]]
            if not outcomes:
                verdict_checks = "none defined"
            elif failed:
                verdict_checks = f"{len(failed)} of {len(outcomes)} failed"
            else:
                verdict_checks = f"{len(outcomes)} passed"

            flagged = [r for r in model_review
                       if isinstance(r, dict) and r.get("form") == form]

            form_rows.append([
                insurer, form, titles.get(form, form), len(fields), filled,
                round(coverage, 3),
                "not reported" if score is None else round(score, 2),
                verdict_checks,
                _verdict(coverage, score, failed, flagged),
                note or "",
            ])

            label_for = dict(zip(fields, labels.get(form, fields)))
            for check in failed:
                review_rows.append([
                    insurer, form, ", ".join(sorted(set(check["fields"]))[:6]), "",
                    "", "Cross-check",
                    f"{check['label']} - got {check['left']:,.1f} vs "
                    f"{check['right']:,.1f}, out by {check['gap']:,.1f}",
                ])
            for flag in flagged:
                field = flag.get("field", "")
                review_rows.append([
                    insurer, form, field, label_for.get(field, ""),
                    values.get(field), "Model",
                    flag.get("reason", ""),
                ])

        for flag in model_review:
            if isinstance(flag, dict) and flag.get("form") not in config:
                review_rows.append([insurer, flag.get("form", "?"),
                                    flag.get("field", ""), "", "", "Model",
                                    flag.get("reason", "")])

    return form_rows, review_rows


def write_summary(wb, results: dict, kind: str, expected_year=None) -> None:
    """Create or replace the summary sheet, and move it to the front."""
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, 0)

    form_rows, review_rows = collect(results, kind, expected_year)

    ws["A1"] = "Extraction summary"
    ws["A1"].font = Font(bold=True, size=13, color="FF17365D")
    ws["A2"] = (f"{'General (NL forms)' if kind == 'general' else 'Life (L forms)'}"
                f"  |  generated {datetime.now():%d %b %Y %H:%M}")
    ws["A2"].font = Font(italic=True, color="FF808080")
    ws["A3"] = ("Verdict is pessimistic by design: 'Review' or 'Check' means open the PDF "
                "to that form. Cross-checks are recomputed arithmetic; model confidence is "
                "self-reported and weaker evidence.")
    ws["A3"].font = Font(italic=True, color="FF808080")

    row = _block(ws, 5, "Per form", FORM_HEADERS, form_rows, verdict_col=9)
    row = _block(ws, row + 2, f"Fields to re-check ({len(review_rows)})",
                 REVIEW_HEADERS, review_rows or [["-", "-", "-", "-", "-", "-",
                                                  "Nothing flagged"]])

    for i, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A7"


def _block(ws, top, title, headers, rows, verdict_col=None):
    """Write a titled header row plus its rows; returns the last row used."""
    ws.cell(row=top, column=1, value=title).font = Font(bold=True, size=11,
                                                        color="FF17365D")
    header_row = top + 1
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for offset, values in enumerate(rows, start=1):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=header_row + offset, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col == len(headers)))
            if col == 6 and isinstance(value, float):
                cell.number_format = "0%"
        if verdict_col:
            cell = ws.cell(row=header_row + offset, column=verdict_col)
            fill = VERDICT_FILLS.get(cell.value)
            if fill:
                cell.fill = fill
                cell.font = Font(bold=True)

    return header_row + len(rows)
