"""Writing extracted data into an output workbook.

Two template shapes are supported, because the two form families are kept in
differently shaped sheets.

  life     Sheets arrive blank. The writer builds everything: row 1 is the
           financial year merged across a block of columns, row 2 the field
           names, column A the insurer. A second year adds a block to the right.

  general  Sheets arrive with a banner, real column headers and number formats
           already in place, one financial year wide. The writer only fills
           cells: columns are fixed, and each insurer takes the next free row.

The general path is the simpler of the two precisely because the template does
more of the work.
"""

import re

import openpyxl

import summary
from config import SHEET_CONFIG
from config_general import HEADER_ROWS, SHEET_CONFIG_GENERAL, SHEET_TITLES

# Data key -> sheet name, for the life templates.
LIFE_SHEETS = {
    "L2": "L2", "L3": "L3", "L4": "L4", "L5": "L5", "L6": "L6", "L7": "L7",
    "L9": "L9", "L22": "L22", "L37": "L37", "L38": "L38",
    "L39_individual": "L39_Individual", "L39_group": "L39_Group",
    "L41": "L41", "L45": "L45",
}


# ------------------------------------------------------------------ life path

def get_or_add_fy_columns(ws, year: str, fields: list, data_start_col: int):
    """
    Row 1 = year label (merged across its fields)
    Row 2 = field subheadings
    Returns the starting column index for this year.
    If year already exists, returns its existing start column.
    """
    for col in range(data_start_col, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == year:
            return col  # already exists

    start_col = data_start_col
    for col in range(data_start_col, ws.max_column + 2):
        if ws.cell(row=1, column=col).value is None:
            if ws.cell(row=2, column=col).value is None:
                start_col = col
                break

    ws.cell(row=1, column=start_col).value = year
    end_col = start_col + len(fields) - 1
    if end_col > start_col:
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1, end_column=end_col
        )

    for i, field in enumerate(fields):
        ws.cell(row=2, column=start_col + i).value = field

    return start_col


def get_or_create_company_row(ws, insurer_name: str, data_start_row: int = 3):
    """Find existing row for insurer or create new one. Starts from row 3."""
    for row in range(data_start_row, ws.max_row + 2):
        val = ws.cell(row=row, column=1).value
        if val == insurer_name:
            return row
        if val is None:
            ws.cell(row=row, column=1).value = insurer_name
            return row


def write_flat_sheet(wb, sheet_name: str, data: dict,
                     insurer_name: str, year: str, fields: list):
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]

    if ws.cell(row=1, column=1).value is None:
        ws.cell(row=1, column=1).value = "Company"

    start_col = get_or_add_fy_columns(ws, year, fields, data_start_col=2)
    company_row = get_or_create_company_row(ws, insurer_name, data_start_row=3)

    for i, field in enumerate(fields):
        col = start_col + i
        ws.cell(row=company_row, column=col).value = data.get(field)


# --------------------------------------------------------------- general path

def write_general_sheet(wb, sheet_title: str, data: dict, insurer_name: str,
                        fields: list, header_rows: int):
    """Fill one pre-headed NL sheet. Columns are fixed; the row is appended.

    Nothing is written above the data: the banner and headers came with the
    template and are better than anything generated here.
    """
    if sheet_title not in wb.sheetnames:
        return

    ws = wb[sheet_title]
    first_data_row = header_rows + 1
    row = get_or_create_company_row(ws, insurer_name, data_start_row=first_data_row)

    for offset, field in enumerate(fields):
        cell = ws.cell(row=row, column=2 + offset)
        cell.value = data.get(field)
        if row > first_data_row:
            # The template only carries formats on its first data row.
            cell.number_format = ws.cell(row=first_data_row,
                                         column=2 + offset).number_format


def template_year(wb) -> str | None:
    """The financial year the general template is laid out for, e.g. 'FY26'.

    Taken from the banner in B1 ("FY26 | NL-1 | Revenue account") so that a
    template rebuilt for a later year needs no code change.
    """
    for title in SHEET_TITLES.values():
        if title not in wb.sheetnames:
            continue
        banner = wb[title].cell(row=1, column=2).value
        match = re.search(r"FY\d{2}", str(banner or ""))
        if match:
            return match.group(0)
    return None


# ------------------------------------------------------------------ entrypoint

def update_excel(results: dict, excel_path: str, kind: str = "life"):
    """Write every successful extraction into the workbook, then summarise it."""
    wb = openpyxl.load_workbook(excel_path)

    expected_year = template_year(wb) if kind == "general" else None

    for insurer_name, result in results.items():
        if result["status"] != "success":
            continue

        data = result["data"]

        if kind == "general":
            for code, fields in SHEET_CONFIG_GENERAL.items():
                values = data.get(code)
                if isinstance(values, dict):
                    title = SHEET_TITLES[code]
                    write_general_sheet(wb, title, values, insurer_name, fields,
                                        HEADER_ROWS.get(title, 2))
        else:
            year = data.get("year", "FY23")
            for data_key, sheet_name in LIFE_SHEETS.items():
                if data_key in data and data_key in SHEET_CONFIG:
                    write_flat_sheet(wb, sheet_name, data[data_key],
                                     insurer_name, year, SHEET_CONFIG[data_key])

    summary.write_summary(wb, results, kind, expected_year)
    wb.save(excel_path)
