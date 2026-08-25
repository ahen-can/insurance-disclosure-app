"""Reading the finished workbook back out for on-screen display.

The copy view is built from the written workbook rather than from the extracted
JSON, so what you copy is exactly what the download would contain: same sheets,
same column order, same values.
"""

import openpyxl

import summary
from config_general import HEADER_ROWS

MAX_HEADER_ROWS = 2      # the templates use "Company / FY" then the field names


def detect_kind(path):
    """Which form family a template is for, from its sheet names.

    Sniffed rather than asked for, so an uploaded workbook needs no extra
    question: only the NL templates name their sheets "NL-something".
    """
    workbook = openpyxl.load_workbook(path, read_only=True)
    names = workbook.sheetnames
    workbook.close()
    return "general" if any(n.startswith("NL-") for n in names) else "life"


def read_workbook(path):
    """[(sheet_name, headers, [row, ...]), ...] with empty sheets dropped."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    out = []
    for name in workbook.sheetnames:
        if name == summary.SHEET_NAME:
            continue          # rendered separately; it is not a data sheet
        worksheet = workbook[name]
        depth = HEADER_ROWS.get(name, MAX_HEADER_ROWS)
        rows = [list(r) for r in worksheet.iter_rows(values_only=True)]
        rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
        if len(rows) <= depth:
            continue                      # header-only sheet, nothing extracted
        headers = _headers(rows, depth)
        data = rows[depth:]
        if data:
            out.append((name, headers, data))
    return out


def _headers(rows, depth=MAX_HEADER_ROWS):
    """The field-name row, falling back to the first row.

    On the deeper NL-34 header the names live on the last header row, with the
    group they belong to on the row above; joining the two keeps "India" and
    "Outside India" distinguishable in the on-screen table.
    """
    candidate = rows[depth - 1] if len(rows) >= depth else rows[0]
    if not any(candidate):
        candidate = rows[0]
    if depth > 2:
        group, last = rows[depth - 2], candidate
        merged, carried = [], ""
        for i, name in enumerate(last):
            carried = group[i] if i < len(group) and group[i] else carried
            merged.append(f"{carried} - {name}" if name and carried else (name or ""))
        return ["" if c is None else str(c) for c in merged]
    return ["" if c is None else str(c) for c in candidate]


def to_tsv(row):
    """One row as tab-separated text, ready to paste across Excel columns."""
    cells = []
    for value in row:
        if value is None:
            cells.append("")
        elif isinstance(value, float) and value.is_integer():
            cells.append(str(int(value)))
        else:
            cells.append(str(value))
    return "\t".join(cells).rstrip("\t")
