"""Offline checks over the template and the write path. No API calls.

    .venv/bin/python tests/test_write.py

The important one is test_column_alignment: values are filed positionally, so a
field list that drifts out of step with the template silently shifts every
number on that sheet into the neighbouring column, and nothing downstream would
notice.
"""

import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

import checks
import summary
from config import SHEET_CONFIG
from config_general import (FIELD_LABELS, HEADER_ROWS, SHEET_CONFIG_GENERAL,
                            SHEET_TITLES)
from excel_writer import update_excel

ROOT = Path(__file__).resolve().parent.parent
GENERAL_TEMPLATE = ROOT / "templates" / "template_general_v1.xlsx"
LIFE_TEMPLATE = ROOT / "templates" / "template_v1.xlsx"

passed = failed = 0


def check(label, condition, detail=""):
    global passed, failed
    if condition:
        passed += 1
    else:
        failed += 1
        print(f"  FAIL {label} {detail}")


def test_column_alignment():
    wb = openpyxl.load_workbook(GENERAL_TEMPLATE)
    check("all sheets present", set(SHEET_TITLES.values()) <= set(wb.sheetnames))
    for code, title in SHEET_TITLES.items():
        ws = wb[title]
        fields = SHEET_CONFIG_GENERAL[code]
        depth = HEADER_ROWS.get(title, 2)
        check(f"{title} width", ws.max_column == 1 + len(fields),
              f"template {ws.max_column - 1} cols, config {len(fields)}")
        headers = [ws.cell(depth, c).value for c in range(2, ws.max_column + 1)]
        check(f"{title} headers", all(h is not None for h in headers))
        check(f"{title} labels", len(FIELD_LABELS[code]) == len(fields))
        check(f"{title} empty", ws.cell(depth + 1, 1).value is None)


def test_check_fields_exist():
    for table, config, name in ((checks.CHECKS_LIFE, SHEET_CONFIG, "life"),
                                (checks.CHECKS_GENERAL, SHEET_CONFIG_GENERAL, "general")):
        for form, rows in table.items():
            check(f"{name} form {form} known", form in config)
            known = set(config.get(form, []))
            for label, left, right in rows:
                for term in list(left) + list(right):
                    if isinstance(term, (int, float)):
                        continue
                    check(f"{name}/{form} field {term}", term.lstrip("-") in known,
                          f"({label})")


def _general_payload(scale=1.0, break_it=False):
    """A self-consistent NL payload: every field 1..n, totals made to add up."""
    data = {"year": "FY26", "company_name": "x", "confidence": {}, "review": []}
    for code, fields in SHEET_CONFIG_GENERAL.items():
        data[code] = {f: round((i + 1) * scale, 2) for i, f in enumerate(fields)}
        data["confidence"][code] = {"score": 0.9, "note": ""}
    nl1 = data["NL1"]
    nl1["total_income"] = sum(nl1[f] for f in
                              ["premiums_earned_net",
                               "profit_loss_on_sale_redemption_of_investments",
                               "interest_dividend_and_rent_gross", "other_income"])
    nl1["total_expenses"] = sum(nl1[f] for f in
                                ["claims_incurred_net", "commission_net",
                                 "operating_expenses", "premium_deficiency"])
    nl1["operating_profit_loss"] = nl1["total_income"] - nl1["total_expenses"]
    if break_it:
        nl1["total_income"] = nl1["total_income"] * 2 + 500
    return data


def test_general_write():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "out.xlsx"
        shutil.copyfile(GENERAL_TEMPLATE, path)
        alpha, beta = _general_payload(1.0), _general_payload(10.0)
        results = {
            "Alpha General Insurance Co. Ltd.": {"status": "success", "data": alpha},
            "Beta General Insurance Co. Ltd.": {"status": "success", "data": beta},
        }
        update_excel(results, str(path), kind="general")

        wb = openpyxl.load_workbook(path)
        for code, title in SHEET_TITLES.items():
            ws = wb[title]
            first = HEADER_ROWS.get(title, 2) + 1
            fields = SHEET_CONFIG_GENERAL[code]
            check(f"{title} row 1 name",
                  ws.cell(first, 1).value == "Alpha General Insurance Co. Ltd.")
            check(f"{title} row 2 name",
                  ws.cell(first + 1, 1).value == "Beta General Insurance Co. Ltd.")
            # Every cell is compared back to the value it was handed, by
            # field name. A column shifted by one shows up here as a whole
            # sheet of mismatches rather than a plausible-looking number.
            for offset, field in enumerate(fields):
                written = ws.cell(first, 2 + offset).value
                check(f"{title}.{field}", written == alpha[code][field],
                      f"wrote {written}, expected {alpha[code][field]}")
            check(f"{title} second row",
                  ws.cell(first + 1, 2).value == beta[code][fields[0]])
            check(f"{title} number format kept",
                  ws.cell(first + 1, 2).number_format != "General")

        check("banner untouched",
              str(wb["NL-1 Revenue Account"]["B1"].value).startswith("FY26"))
        check("summary sheet first", wb.sheetnames[0] == summary.SHEET_NAME)


def test_summary_flags_a_broken_total():
    good = _general_payload(1.0)
    bad = _general_payload(1.0, break_it=True)
    bad["review"] = [{"form": "NL4", "field": "motor_od",
                      "reason": "two candidate figures"}]
    results = {
        "Clean Co": {"status": "success", "data": good},
        "Broken Co": {"status": "success", "data": bad},
        "Failed Co": {"status": "error", "error": "timeout"},
    }
    form_rows, review_rows = summary.collect(results, "general")

    nl1_broken = [r for r in form_rows if r[0] == "Broken Co" and r[1] == "NL1"]
    check("broken NL1 verdict", nl1_broken and nl1_broken[0][8] == "Check",
          nl1_broken[0][8] if nl1_broken else "missing")
    nl1_clean = [r for r in form_rows if r[0] == "Clean Co" and r[1] == "NL1"]
    check("clean NL1 verdict", nl1_clean and nl1_clean[0][8] == "OK",
          nl1_clean[0][8] if nl1_clean else "missing")
    check("failure row present",
          any(r[0] == "Failed Co" and r[8] == "Check" for r in form_rows))
    check("cross-check raised",
          any(r[5] == "Cross-check" and r[1] == "NL1" for r in review_rows))
    check("model flag carried",
          any(r[5] == "Model" and r[2] == "motor_od" for r in review_rows))
    # The fixture only makes NL1 internally consistent, so the other forms
    # are expected to fail their own checks -- NL1 is the controlled comparison.
    check("clean NL1 not flagged",
          not any(r[0] == "Clean Co" and r[1] == "NL1" for r in review_rows))


def test_year_mismatch_flagged():
    data = _general_payload(1.0)
    data["year"] = "FY25"
    rows, review = summary.collect({"A": {"status": "success", "data": data}},
                                   "general", expected_year="FY26")
    check("year mismatch flagged",
          any(r[5] == "Year check" for r in review))
    rows, review = summary.collect({"A": {"status": "success", "data": data}},
                                   "general", expected_year="FY25")
    check("matching year not flagged",
          not any(r[5] == "Year check" for r in review))


def test_life_path_still_works():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "life.xlsx"
        shutil.copyfile(LIFE_TEMPLATE, path)
        data = {"year": "FY26"}
        for form, fields in SHEET_CONFIG.items():
            data[form] = {f: i + 1 for i, f in enumerate(fields)}
        update_excel({"Alpha Life Insurance Limited":
                      {"status": "success", "data": data}}, str(path), kind="life")

        wb = openpyxl.load_workbook(path)
        ws = wb["L2"]
        check("life year label", ws.cell(1, 2).value == "FY26")
        # Header is now the human-readable label, not the raw JSON key -- this
        # is the change the user asked for, so it is what the test pins down.
        check("life field header", ws.cell(2, 2).value == "Amounts transferred",
              ws.cell(2, 2).value)
        check("life insurer", ws.cell(3, 1).value == "Alpha Life Insurance Limited")
        check("life first value", ws.cell(3, 2).value == 1)
        check("life Indian grouping", ws.cell(3, 2).number_format == "#,##,##0.##;[Red](#,##,##0.##)",
              ws.cell(3, 2).number_format)
        check("life summary written", summary.SHEET_NAME in wb.sheetnames)

        rows, review = summary.collect(
            {"Alpha Life Insurance Limited": {"status": "success", "data": data}},
            "life")
        l2_row = [r for r in rows if r[1] == "L2"]
        check("life summary uses form code", bool(l2_row))


def test_collect_usage():
    results = {
        "Alpha": {"status": "success", "data": {
            "usage": {"input_tokens": 10000, "output_tokens": 2000, "total_tokens": 12000}}},
        "Beta": {"status": "success", "data": {
            "usage": {"input_tokens": 5000, "output_tokens": 1000, "total_tokens": 6000}}},
        "Gamma (no usage reported)": {"status": "success", "data": {}},
        "Delta (failed)": {"status": "error", "error": "timeout"},
    }
    rows = summary.collect_usage(results)
    check("usage rows: two reporting + one total", len(rows) == 3, rows)
    check("Gamma excluded, not zeroed",
          not any(r[0] == "Gamma (no usage reported)" for r in rows))
    check("Delta excluded", not any(r[0] == "Delta (failed)" for r in rows))
    total = rows[-1]
    check("total row label", total[0] == "Total")
    check("total input summed", total[1] == 15000, total[1])
    check("total output summed", total[2] == 3000, total[2])
    check("total tokens summed", total[3] == 18000, total[3])

    check("no usage at all -> empty", summary.collect_usage(
        {"X": {"status": "success", "data": {}}}) == [])


def test_usage_reaches_the_workbook():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "usage.xlsx"
        shutil.copyfile(GENERAL_TEMPLATE, path)
        data = _general_payload(1.0)
        data["usage"] = {"input_tokens": 43210, "output_tokens": 6789, "total_tokens": 49999}
        update_excel({"Usage Co": {"status": "success", "data": data}}, str(path),
                     kind="general")
        wb = openpyxl.load_workbook(path)
        ws = wb[summary.SHEET_NAME]
        values = [c.value for row in ws.iter_rows() for c in row]
        check("usage title on sheet", "Gemini token usage" in values)
        check("input tokens on sheet", 43210 in values)
        check("output tokens on sheet", 6789 in values)


for test in [test_column_alignment, test_check_fields_exist, test_general_write,
             test_summary_flags_a_broken_total, test_year_mismatch_flagged,
             test_life_path_still_works, test_collect_usage,
             test_usage_reaches_the_workbook]:
    test()

print(f"{passed} passed, {failed} failed")
sys.exit(1 if failed else 0)
