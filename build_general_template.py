"""Turn the hand-made general-insurer workbook into a shippable template.

`c.xlsx` was laid out by hand with all 27 general insurers pre-seeded down
column A. The extractor appends one row per PDF instead, so the seeded names
have to go -- but everything above them (the banner, the headers, the number
formats) is worth keeping, which is why this is a transform rather than a
rebuild from scratch.

One blank data row is deliberately left behind on every sheet. It carries the
number formats, so the writer has something to copy when it adds further rows,
and its empty column A is what `get_or_create_company_row` fills first.

    python build_general_template.py [source.xlsx]
"""

import shutil
import sys
from pathlib import Path

import openpyxl

from config_general import HEADER_ROWS, SHEET_TITLES, SHEET_CONFIG_GENERAL

DEFAULT_SOURCE = Path.home() / "Desktop" / "c.xlsx"
TARGET = Path(__file__).parent / "templates" / "template_general_v1.xlsx"


def build(source: Path, target: Path) -> None:
    shutil.copyfile(source, target)
    workbook = openpyxl.load_workbook(target)

    expected = set(SHEET_TITLES.values())
    missing = expected - set(workbook.sheetnames)
    if missing:
        raise SystemExit(f"{source} is missing sheets: {sorted(missing)}")

    for title in SHEET_TITLES.values():
        worksheet = workbook[title]
        header_rows = HEADER_ROWS.get(title, 2)
        first_data_row = header_rows + 1

        # Keep the first data row for its number formats; drop the rest.
        surplus = worksheet.max_row - first_data_row
        if surplus > 0:
            worksheet.delete_rows(first_data_row + 1, surplus)
        worksheet.cell(row=first_data_row, column=1).value = None

        # c.xlsx declares a few columns past the last header (NL-18 stops at
        # G but claims R, NL-41 claims a blank T). Left in place they show up
        # as empty trailing columns in the results table.
        used = 1 + len(SHEET_CONFIG_GENERAL[_code_for(title)])
        if worksheet.max_column > used:
            worksheet.delete_cols(used + 1, worksheet.max_column - used)

        worksheet.freeze_panes = worksheet.cell(row=first_data_row, column=2)

    workbook.save(target)
    print(f"Wrote {target} ({len(SHEET_TITLES)} sheets)")


def _code_for(title: str) -> str:
    for code, sheet_title in SHEET_TITLES.items():
        if sheet_title == title:
            return code
    raise KeyError(title)


if __name__ == "__main__":
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    if not source.exists():
        raise SystemExit(f"Source workbook not found: {source}")
    build(source, TARGET)
